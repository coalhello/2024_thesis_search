from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook

def get_info(value):
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/subject.txt", "r") as sub:
        info = sub.readline()

    count = 1
    webdriver_path = 'chromedriver.exe'

    chrome_options = Options()
    chrome_options.add_argument('--headless')

    service = Service(webdriver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    url = 'https://www.dbpia.co.kr/search/topSearch?searchOption=all&query=' + info
    driver.get(url)

    titleL = []
    timeL = []
    authorL = []

    try:
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'searchResultList')))
        thesis_titles = driver.find_elements(By.CLASS_NAME, 'thesis__tit')
        thesis_item = driver.find_elements(By.CSS_SELECTOR, 'span.thesis__item:not(.thesis__record):not(.thesis__useCount):not([m-hidden])')
        thesis_author = driver.find_elements(By.CSS_SELECTOR, 'div.thesis__item')
        for index, title in enumerate(thesis_titles):
            if index >= value:
                break
            titleL.append(title.text)
        for index, time in enumerate(thesis_item):
            if index >= value:
                break
            timeL.append(time.text)
        for index, author in enumerate(thesis_author):
            if index >= value:
                break
            authorL.append(author.text)
    except:
        print("짧은 검색 오류")
    finally:
        driver.quit()

    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/title.txt", "w+") as safe_title:
        safe_title.write('\n'.join(titleL))
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/time.txt", "w+") as safe_time:
        safe_time.write('\n'.join(timeL))
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/author.txt", "w+") as safe_author:
        safe_author.write('\n'.join(authorL))

def get_max():
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/subject.txt", "r") as sub:
        info = sub.readline()

    webdriver_path = 'chromedriver.exe'

    chrome_options = Options()
    chrome_options.add_argument('--headless')

    service = Service(webdriver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    url = 'https://www.dbpia.co.kr/search/topSearch?searchOption=all&query=' + info
    driver.get(url)

    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'totalCount')))
        mxv = driver.find_element(By.ID, 'totalCount')
        return mxv.text.strip()
    except:
        print("주제 검색 오류")
    finally:
        driver.quit()

def make_yxl():
    L_time = []
    L_title = []
    L_author = []

    wb = Workbook()
    file_name = 'C:/Users/siri4/OneDrive/바탕 화면/Py_P/액셀/thesis.xlsx'
    ws = wb.active
    ws.title = "논문 정리"

    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/time.txt") as time:
        for line in time:
            L_time.append(line.strip())
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/author.txt") as author:
        for line in author:
            L_author.append(line.strip())
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/title.txt") as title:
        for line in title:
            L_title.append(line.strip())

    for i in range(len(L_time)):
        ws.cell(row=i + 1, column=1, value=L_time[i])
        ws.cell(row=i + 1, column=2, value=L_author[i])
        ws.cell(row=i + 1, column=3, value=L_title[i])

    wb.save(file_name)
    wb.close()

def move_link(item):
    L_title = []

    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/subject.txt", "r") as sub:
        info = sub.readline()
    with open("C:/Users/siri4/OneDrive/바탕 화면/Py_P/storage/title.txt") as title:
        for line in title:
            L_title.append(line.strip())

    webdriver_path = 'chromedriver.exe'

    chrome_options = Options()
    chrome_options.add_argument('--headless')

    service = Service(webdriver_path)
    driver = webdriver.Chrome(service=service, options=chrome_options)

    url = 'https://www.dbpia.co.kr/search/topSearch?searchOption=all&query=' + info
    driver.get(url)
    try:
        buttons = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "thesis__pdfBtn")))
        id_list = [button.get_attribute('id') for button in buttons]

        temp = L_title.index(item)
        user_thesis = id_list[temp]

        button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, user_thesis)))
        button.click()

        handles = driver.window_handles
        new_window_handle = None
        for handle in handles:
            if handle != driver.current_window_handle:
                new_window_handle = handle
                break

        if new_window_handle:
            driver.switch_to.window(new_window_handle)
            new_url = driver.current_url
            return new_url
    except:
        print("링크 이동 오류")
    finally:
        driver.quit()